import os
from http.client import responses

import pandas as pd
from datetime import datetime

from flask import jsonify
from openpyxl import Workbook, load_workbook

from swagger_server.models import ResponseError


class FilesMethods:
    FORMATO = [
        "CUENTA TITAN", "Referencia interna", "DOCUMENTO IDENTIFICACION", "NOMBRE DEL CLIENTE", "TELEFONOS", "CIUDAD",
        "ESTADO CUENTA", "TIPO CUENTA", "TIPO DE NEGOCIO", "FORMA DE PAGO", "TRANSACCION", "NOM_TRANSACCION",
        "# FAC PEN CARGA", "# FACTURAS PENDIENTE", "SALDO ORIGINAL VENC", "GESTION COBRANZA TOTAL",
        "TOTAL A PAGAR VENCIDO", "SALDO ACTUAL", "TOTAL A PAGAR", "MOVIMIENTOS (+)", "MOVIMIENTOS (-)", "TOTAL PAGO",
        "Valor ajuste", "ESTADO_LIQUIDACION", "LIQ. GC POR VALIDAR", "Gestion OK GC_NO GC", "Fecha pago",
        "[RESUMEN CONTACTO IVR]", "Fecha IVR", "[RESUMEN CONTACTO LLAMADA]", "Fecha llamada", "[LIQ. TVCABLE]",
        "CONVENIO", "Respaldo", "CORREO CLIENTE", "EMAIL_CAMPAÑA", "CELULAR CAMPAÑA", "Fecha terminacion", "Empresa",
        "Dias Vencidos", "NOMBRE_ARCHIVO"
    ]

    @staticmethod
    def fusion_files(archivos):
        """Fusiona múltiples archivos Excel en uno solo, respetando su estructura y añadiendo una columna 'NOMBRE_ARCHIVO'."""
        # Ruta donde se almacenará el archivo fusionado
        today = datetime.now().strftime("%Y-%m-%d")
        file_name = f"fusion_Cobranza_{today}.xlsx"
        save_path = os.path.join("swagger_server", "files", file_name)

        # Verificar si el archivo ya existe
        if not os.path.exists(save_path):
            # Crear un nuevo archivo con las cabeceras definidas en FORMATO
            wb = Workbook()
            ws = wb.active
            ws.title = "Fusión Cobranza"
            ws.append(FilesMethods.FORMATO)  # Agregar las cabeceras
            wb.save(save_path)
        # Cargar el archivo existente para agregar los datos
        wb = load_workbook(save_path)
        ws = wb.active
        # Iterar sobre los archivos proporcionados y procesar los datos
        for archivo in archivos:
            try:
                # Leer el archivo Excel o CSV
                if archivo.filename.endswith('.csv'):
                    df = pd.read_csv(archivo)
                else:
                    df = pd.read_excel(archivo, engine='openpyxl')
                # Añadir una nueva columna con el nombre del archivo
                df["NOMBRE_ARCHIVO"] = archivo.filename
                # Escribir los datos en el archivo fusionado (omitimos los encabezados de los archivos originales)
                for _, row in df.iterrows():
                    ws.append(row.tolist())
            except Exception as e:
                response = ResponseError(
                    error_code= 500,
                    message=f'Error Detectado: {e}',
                    internal_transaction_id='099',
                    external_transaction_id='099'
                )
                return jsonify(response), 500

        # Guardar el archivo fusionado
        wb.save(save_path)

        # Retornar el nombre del archivo creado
        return file_name
